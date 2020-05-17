import os
import requests
import time
import platform
import re

from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.keys import Keys

import openpyxl
from openpyxl import load_workbook

from nsQtUtil import *
from PyQt5.QtWidgets import *

from CrawlData import CrawlData

# 네이버 쇼핑
URL_NAVER_SHOP = "https://shopping.naver.com/"

# 저장 폴더
FOLDER_PATH = "SHOP CRAWLER"

# 모든 카테고리 정보
bigCategoryTextList = []
categoryTextList = [[]]
categoryLinkList = [[]]

# 저장할 엑셀 정보
#gExcelFileName = ''

# 크롤링한 데이터
crawlDataList = ''#[[[] for first in range(100)] for second in range(50)]

# webDriver . Chrome driver Ver.80
if platform.system() == "Windows":
    driver = webdriver.Chrome("chromedriver.exe")
elif platform.system() == "Darwin":
    driver = webdriver.Chrome("/Users/igyutae/Documents/GitRepository/NaverShoppingCrawler/NaverShoppingCrawler/NaverShoppingCrawler/chromedriver")

# 카테고리 정보들 가져오기
def GetAllCategoryLink():
    driver.get(URL_NAVER_SHOP)
    time.sleep(1)
    
    # 카테고리 리스트 요소
    div_elems = driver.find_elements_by_xpath("//ul[@class='co_category_list']/li[@*]/a[@href='#']")
    
    CrawlBigCategory(driver.page_source)

    categoryIndex = 0
    for element in div_elems:
        element.send_keys(Keys.ENTER)
        CrawlAllCategory(driver.page_source, categoryIndex)
        categoryIndex += 1

def CrawlBigCategory(html):    
    bsObj = BeautifulSoup(html, "html.parser")    
    itemList = bsObj.find_all('ul', {"class" :"co_category_list"})
    
    if bsObj == None:
        print("아무것도 없는 카테고리 입니다. (url : ", url, ")")
        return

    for link in itemList[1].find_all('a'):
        if hasattr(link, 'attrs') and 'href' in link.attrs:
            bBigCategoryCheck  = "#" in link.attrs['href']
            if bBigCategoryCheck == True:
                bigCategoryTextList.append(link.text.strip())
                continue

def CrawlAllCategory(html, categoryIndex):
    bsObj = BeautifulSoup(html, "html.parser")    
    itemList = bsObj.find_all('ul', {"class" :"co_category_list"})
    
    if bsObj == None:
        print("ERROR : URL FAULT [001] (url : ", url, ")")
        return

    categoryTextList.append([])
    categoryLinkList.append([])

    #?? 왜 1만 되나
    for link in itemList[1].find_all('div', {"class" : "co_position"}):            
        for httpLink in link.find_all('a'):
            if hasattr(httpLink, 'attrs') and 'href' in httpLink.attrs:
                bHttpCheck = "http" in httpLink.attrs['href']
                if bHttpCheck == True:
                    categoryTextList[categoryIndex].append(httpLink.text.strip())
                    categoryLinkList[categoryIndex].append(httpLink.attrs['href'].strip())
                    continue
            
def ClickTab(xpath):
    if xpath == None:
        print("ERROR XPATH NULL [007]")
        return False
    
    try:
        element = driver.find_element_by_xpath(xpath)
        element.send_keys(Keys.ENTER)
        time.sleep(1)

        return True
    except:
        print("ERROR element NULL [008] : " + xpath)
        return False
    

# 모든 아이템의 정보 가져오기
def CrawlItemInfo(url, fileName, pageCount, listIndex0, listIndex1):
    try:
        driver.get(url)
    except:
        print("ERROR url ERROR [009] : " + url)
        return

    # 가격 비교 탭이 있는지 검사
    rqResult = requests.get(driver.current_url)
    bsObj = BeautifulSoup(rqResult.content, "html.parser")

    if bsObj == None:
        print("ERROR : URL FAULT [009] (url : ", driver.current_url, ")")
        return

    checkCategory = bsObj.find('li', {"class" : "snb_compare"})

    if checkCategory == None:
        print("해당 카테고리에는 등록된 상품이 하나도 없습니다.")
        return

    # 가격 비교 탭으로 이동
    result = ClickTab("//ul[@class='snb_list']/li[@class='snb_compare']/a[@href='#']")

    if result == False:
        return

    tabLink = driver.current_url
    rqResult = requests.get(tabLink)
    bsObj = BeautifulSoup(rqResult.content, "html.parser")

    if bsObj == None:
        print("ERROR : URL FAULT [015] (url : ", tabLink, ")")
        return

    # 페이지 버튼 검색
    pageButton = bsObj.find('div', {"class" : "sort_content"})

    if pageButton == None:
        time.sleep(5)
        pageButton = bsObj.find('div', {"class" : "sort_content"})
        if pageButton == None:
            print("ERROR : pageButton FAULT [016] (url : ", tabLink, ")")
            return

    if pageButton.find('div', {"class" : "search_none"}) != None:
        print("가격을 비교할 수 있는 제품이 없습니다. (url : ", tabLink, ")")
        return
    elif pageButton.find('div', {"class" : "co_paginate"}) != None:
        pageButton = pageButton.find('div', {"class" : "co_paginate"})

        # 페이지가 한 페이지 뿐임
        if pageButton.find('a', {"href" : "#"}) == None:
            print("해당 카테고리는 페이지가 한 개 뿐입니다. (url : ", tabLink, ")")
        else:
            # 두번째 탭 클릭
            ClickTab("//div[@class='co_paginate']/a[@href='#']")
            tabLink = driver.current_url
            # 기본 페이지 Index 설정 (1페이지로)
            tabLink = tabLink.replace("2", "1", 1)

    # 1페이지 정보 가져오기
    rqResult = requests.get(tabLink)
    bsObj = BeautifulSoup(rqResult.content, "html.parser")
    
    if bsObj == None:
        print("ERROR : URL FAULT [002] (url : ", tabLink, ")")
        return

    # 크롤링 데이터 생성
    crawlData = CrawlData()

    contentIndex = 0

    # 요청한 페이지만큼 크롤링
    for pageNumber in range(1, pageCount + 1):
        itemList = bsObj.find('ul', {"class" : "goods_list"})

        if itemList == None:
            print("ERROR : URL FAULT [003] (url : ", tabLink, ")")
            break

        for content in itemList.contents:
            if hasattr(content, 'attrs') and 'class' in content.attrs:
                # 광고 거르기
                bAdCheck = 'ad' in content.attrs['class']
                if bAdCheck == False:
                    # INFO
                    itemInfo = content.find('div', {"class" : "info"})
                    itemTitle = itemInfo.find('div', {"class" : "tit"})
                    itemTitle = itemTitle.find('a', {"class" : "link"})
                    itemPrice = itemInfo.find('span', {"class" : "price"})
                    itemEtc = itemInfo.find('span', {"class" : "etc"})                
                    
                    # ETC
                    itemReview = itemEtc.find("em")
                    itemDate = itemEtc.find('span', {"class" : "date"})

                    crawlData.itemDataList.append([])
                    
                    crawlData.itemDataList[contentIndex].append(itemTitle.text.strip())

                    if itemPrice.text.strip() != "판매중단":
                        itemPriceText = (itemPrice.text[0:itemPrice.text.find("원")])
                        itemPriceText = itemPriceText.replace("최저", "").replace("모바일", "").replace("가격", "").replace(",", "").strip()
    
                        if itemPriceText == "":
                            itemPriceText = "-"
                    else:
                        itemPriceText = itemPrice.text.strip()

                    crawlData.itemDataList[contentIndex].append(itemPriceText)

                    itemSellerCountText = (itemPrice.text[itemPrice.text.find("원") + 1:])
                    itemSellerCountText = itemSellerCountText.replace("판매처", "").replace(",", "").replace("QR코드", "").strip()

                    if itemSellerCountText == "":
                        itemSellerCountText = "-"

                    crawlData.itemDataList[contentIndex].append(itemSellerCountText)

                    itemReviewCountText = itemReview.text.replace(",", "").strip()

                    if itemReviewCountText == "":
                        itemReviewCountText = "-"

                    crawlData.itemDataList[contentIndex].append(itemReviewCountText)

                    itemDateText = itemDate.text.replace("등록일", "").strip()
                    crawlData.itemDataList[contentIndex].append(itemDateText)
                    
                    # 판매처가 여러개 있을 경우 판매처 정보 크롤링
                    for itemPriceContent in itemPrice.contents:
                        if hasattr(itemPriceContent, 'attrs'):
                            bPriceCheck = 'href' in itemPriceContent.attrs
                            if bPriceCheck == True:
                                itemSellLink = itemPriceContent.attrs['href'].strip()
                                CrawlDetailItemInfo(itemSellLink, crawlData, contentIndex, itemTitle.text.strip())
                                time.sleep(1)                                
                    
                    contentIndex += 1

        
        # 다음 페이지로 이동
        tabLink = tabLink.replace(str(pageNumber), str(pageNumber + 1), 1)
        rqResult = requests.get(tabLink)
        bsObj = BeautifulSoup(rqResult.content, "html.parser")

        if bsObj == None:
            print("페이지가 입력하신 기준 페이지보다 적습니다. [010] (url : ", tabLink, ")")
            break

    SaveItemListAsExcelEx(crawlData, fileName, listIndex0, listIndex1)

    #print("세부 카테고리 - \"" , categoryTextList[listIndex0][listIndex1], "\" : 크롤링 데이터 추가")
    #crawlDataList[listIndex0][listIndex1].append(crawlData)
               
# 여러 판매처의 아이템 정보 가져오기
def CrawlDetailItemInfo(url, crawlData, contentIndex, itemTitle):    
    rqResult = requests.get(url)
    bsObj = BeautifulSoup(rqResult.content, "html.parser")

    ##
    # INFO INNER
    infoInner = bsObj.find('div', {"class" : "info_inner"})

    if infoInner == None:
        print("ERROR INFO INNER NONE [019] (아이템 이름 : ", itemTitle, ")")
        return False

    infoText = infoInner.text.strip()
    infoText = infoText[:infoText.find("등록일")]
    infoTextList = infoText.split("\n")

    while '' in infoTextList:
        infoTextList.remove('')

    infoTextListLength = len(infoTextList)

    # 비어있는 경우들은 가데이터를 넣어준다
    if infoTextListLength == 0:
        crawlData.itemDataList[contentIndex].append("-")
        crawlData.itemDataList[contentIndex].append("-")
    elif infoTextListLength == 1:
        crawlData.itemDataList[contentIndex].append("-")

    for index in range(0,len(infoTextList)):
        if infoTextList[index] == "":
            continue

        infoTextList[index] = infoTextList[index].replace("제조사", "")
        infoTextList[index] = infoTextList[index].replace("브랜드", "")
        infoTextList[index] = infoTextList[index].strip()

        crawlData.itemDataList[contentIndex].append(infoTextList[index])

    ##
    # TABLE LIST
    tableList = bsObj.find('table', {"class" : "tbl_lst"})

    if tableList == None:
        print("ERROR TABLE LIST NONE [013] (아이템 이름 : ", itemTitle, ")")
        return False

    itemList = tableList.find_all('tr', {"class" : "_itemSection"})
    
    for item in itemList:
        if hasattr(item, 'attrs') and 'class' in item.attrs:
            # PRICE
            price = item.find('td', {"class" : "price"})
            priceText = price.text.strip()

            # 첫 번째가 인기인 경우가 있으면 두 번째가 최저이다.
            # 하나밖에 없으면 인기여도 쓴다
            if len(itemList) != 1 and ("인기" in priceText) == True:
                continue                        

            # 링크
            mall = item.find('span', {"class" : "mall"})
            if hasattr(mall.contents[0], 'attrs'):
                bCheck = 'href' in mall.contents[0].attrs
                if bCheck == True:
                    itemSellLink = mall.contents[0].attrs['href'].strip()
                    crawlData.itemDataList[contentIndex].append(itemSellLink)
                
                else:
                    crawlData.itemDataList[contentIndex].append("-")
            else:
                    crawlData.itemDataList[contentIndex].append("-")

            # 배송비
            gift = item.find('td', {"class" : "gift"})
            if gift != None:
                if gift.text.strip() != "":
                    crawlData.itemDataList[contentIndex].append(gift.text.strip())
                else:
                    crawlData.itemDataList[contentIndex].append("-")
            else:                
                crawlData.itemDataList[contentIndex].append("-")

            # 부가정보
            info = item.find('td', {"class" : "info"})
            if info != None:
                if info.text.strip() != "":
                    crawlData.itemDataList[contentIndex].append(info.text.strip())
                else:
                    crawlData.itemDataList[contentIndex].append("-")
            else:
                crawlData.itemDataList[contentIndex].append("-")

            break

    return True

##
##
##
# 폴더가 없으면 생성한다
def CheckAndCreateFolder(folderPath):
    try:
        if not os.path.exists(folderPath):
            os.makedirs(folderPath)
    except OSError:
        print ("Error: Creating folder [004]: " + folderPath)


# 엑셀 생성
def CreateExcelFile(fileName):
    workBook = openpyxl.Workbook()
    workBook.save(FOLDER_PATH + "/" + fileName + ".xlsx")
    workBook.close()

# 엑셀 매번 저장
def SaveItemListAsExcelEx(crawlData, fileName, index0, index1):
    workBook = load_workbook(FOLDER_PATH + "/" + fileName + ".xlsx")
    
    sheetTitle = categoryTextList[index0][index1].replace('/', '.')
    newSheet = workBook.create_sheet(sheetTitle)

    newSheet.append(["제품 이름", "제품 가격", "판매처", "리뷰", "제품 등록일", "제조사", "브랜드", "링크", "배송비", "기타"])
    
    for index in range(0, len(crawlData.itemDataList)):
        newSheet.append(crawlData.itemDataList[index])
        
    workBook.save(FOLDER_PATH + "/" + fileName + ".xlsx")
    workBook.close()

    print("크롤링 완료 : "  , bigCategoryTextList[index0]  , " - " , categoryTextList[index0][index1])


# 액셀 저장
def SaveItemListAsExcel(fileName):
    workBook = openpyxl.Workbook()

    excelSheet = []
    
    excelSheetIndex = 0

    for index0 in range(0, len(crawlDataList)):
        for index1 in range(0, len(crawlDataList[index0])):
            for index2 in range(0, len(crawlDataList[index0][index1])):
                print("크롤링 완료 : "  , bigCategoryTextList[index0]  , " - " , categoryTextList[index0][index1])
                sheetTitle = categoryTextList[index0][index1].replace('/', '.')
                newSheet = workBook.create_sheet(sheetTitle)

                excelSheet.append(newSheet)
                excelSheet[excelSheetIndex].append(["제품 이름", "제품 가격", "판매처", "리뷰", "제품 등록일", "제조사", "브랜드", "링크", "배송비", "기타"])

                for index3 in range(0, len(crawlDataList[index0][index1][index2].itemDataList)):
                    excelSheet[excelSheetIndex].append(crawlDataList[index0][index1][index2].itemDataList[index3])

                excelSheetIndex += 1

    CheckAndCreateFolder(FOLDER_PATH)

    workBook.save(FOLDER_PATH + "/" + fileName + ".xlsx")
    workBook.close()

    print("[저장 완료]")

def CheckBoolList(boolList):
    for index in range(0, len(boolList)):
        if boolList[index] == True:
            return True
    return False

# pageCount : 몇 페이지까지 크롤링 할 것인지
# boolList  : GUI에서 체크한 정보들만 boolList[index0][index1] = True
def StartCrawling(pageCount, excelFileName, boolList):
    if pageCount == None:
        pageCount = 1
        print("NO PAGE COUNT [005]")

    if excelFileName == '' or excelFileName == None:
        excelFileName = "NAVER SHOP DATA"
        print("NO EXCEL FILE NAME [017]")

    if boolList == None:
        boolList = [True for b in range(100)]
        print("NO BOOL LIST [006]")

    CheckAndCreateFolder(FOLDER_PATH)
    CreateExcelFile(excelFileName)
        
    # 각 링크에서 데이터 크롤링
    for index0 in range(0, len(bigCategoryTextList)):
        if CheckBoolList(boolList[index0]) == False:
            print("카테고리 - \"", bigCategoryTextList[index0], "\" : 비어 있음")
            continue

        if index0 != 0:
            print("카테고리 - \"", bigCategoryTextList[index0], "\" : 3초 대기")
            time.sleep(3)
            
        for index1 in range(0, len(categoryLinkList[index0])):
            if boolList[index0][index1] == False:
                continue

            if index1 != 0:
                print("세부 카테고리 - \"", categoryTextList[index0][index1], "\" : 1초 대기")
                time.sleep(1)
            
            CrawlItemInfo(categoryLinkList[index0][index1], excelFileName, pageCount, index0, index1)
        
        print("----------------------------------------------------")
       
    print("크롤링 종료 : " + excelFileName + " 저장 완료")
    #SaveItemListAsExcel(excelFileName)

##
##
##
def OpenWindow(window):
    screenWidth = 1400
    screenHeight = 480

    # Window 크기 및 위치 설정
    myDesktop = QApplication.desktop()
    rect = myDesktop.screenGeometry()
    window.setGeometry(rect.width() / 2 - screenWidth / 2, rect.height() / 2 - screenHeight / 2, screenWidth, screenHeight)
    
    window.InitializeWindow(bigCategoryTextList, categoryTextList, StartCrawling)
    
##
##
##
def app_init(window):
    window.setWindowTitle("Naver Shopping Crawler (ver.1.0)")
    
    ##
    # 카테고리 링크 가져오기
    GetAllCategoryLink()
      
    ##
    # 윈도우 설정
    maxIndex1Count = 0
    for i in range(0, len(categoryLinkList)):
        if maxIndex1Count > len(categoryLinkList[i]):
            maxIndex1Count = len(categoryLinkList[i])

    maxIndex1Count += 1

    OpenWindow(window)


def main():    
    ##GUI CODE
    app = QApplication(sys.argv)
    window = MainWindow()
    app_init(window)
    
    widget = window.centralWidget()

    window.show()
    app.exec_()

##
##
##
#MAIN
main()